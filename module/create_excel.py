"""
create_excel.py

dictionary 데이터를 전송 받아서 엑셀에 입력합니다
"""


import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill


from module.create_log import logging
from module.data import committee_dict, organization_dict, person_dict


def create_excel(excel_list, excel_path, bookmark_list, output_path):
    """받은 데이터를 토대로 엑셀을 생성합니다"""
    wb = _load_excel(excel_path)
    ws = wb.active

    wb2 = _load_excel(os.path.join(output_path, 'total_list.xlsx'), 2)
    ws2 = wb2.active

    last_row = 2

    for item in excel_list:
        ws.cell(row=last_row, column=1, value=ws.max_row)
        ws.cell(row=last_row, column=2, value=item['org'])
        ws.cell(row=last_row, column=3,
                value=organization_dict[item['org']] if item['org']
                in organization_dict else None)
        ws.cell(row=last_row, column=4, value=item['cmt'])
        ws.cell(row=last_row, column=5,
                value=committee_dict[item['cmt']] if item['cmt']
                in committee_dict else None)
        ws.cell(row=last_row, column=6, value=item['name'])
        ws.cell(row=last_row, column=7,
                value=person_dict[item['name']] if item['name'] in person_dict else None)
        ws.cell(row=last_row, column=8, value=1)
        last_row += 1

        cmt = item['cmt']
        org = item['org']

    last_row = 2
    ws2_last_row = ws2.max_row + 1

    for bookmark in bookmark_list:
        ws.cell(row=last_row, column=9,
                value=bookmark['split_bookmark_name'])
        ws.cell(row=last_row, column=10,
                value=f"{bookmark['split_pdf_name']}.PDF")
        ws2.cell(row=ws2_last_row, column=1, value=cmt)
        ws2.cell(row=ws2_last_row, column=2, value=org)
        ws2.cell(row=ws2_last_row, column=3, value=f"{
                 bookmark['split_pdf_name']}.PDF")
        ws2.cell(row=ws2_last_row, column=4, value=bookmark['real_file_name'])

        last_row += 1
        ws2_last_row += 1
        filename = bookmark['real_file_name']

    ws2.cell(row=ws2_last_row, column=1, value=cmt)
    ws2.cell(row=ws2_last_row, column=2, value=org)
    ws2.cell(row=ws2_last_row, column=3, value=os.path.basename(excel_path))
    ws2.cell(row=ws2_last_row, column=4, value=filename)

    wb.save(excel_path)
    wb2.save(os.path.join(output_path, 'total_list.xlsx'))


def _has_header(wb, path, select=1):
    """엑셀 header가 존재하는지 확인합니다. 존재하지 않을 경우 새로 생성합니다"""
    ws = wb.active
    first_row = ws[1]
    header_exists = any(cell.value for cell in first_row)
    headers = []

    if not header_exists:
        if select == 2:
            headers = ['위원회', '피감기관', '답변 책자 파일명', '파일명']
        else:
            headers = ['일련번호', '기관명', '기관코드', '위원회명', '위원회 코드',
                       '위원(의원)명', '위원(의원) 코드', '질의유형', '질의', '답변 책자 파일명']

        fill_color = PatternFill(start_color='4f81bd',
                                 end_color='4f81bd', fill_type='solid')

        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
            ws.cell(row=1, column=col_idx).fill = fill_color

    wb.save(path)

    return wb


def _load_excel(excel_path, select=1):
    """엑셀을 불러옵니다. 파일이 없는 경우 새로 생성됩니다"""
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
    else:
        try:
            if not os.path.exists(os.path.dirname(excel_path)):
                os.makedirs(os.path.dirname(excel_path))
            wb = Workbook()
            wb.save(excel_path)
            wb = load_workbook(excel_path)
        except Exception as e:  # pylint: disable=W0703
            e = "엑셀 파일 생성 오류"
            logging(e, '', os.path.dirname(excel_path))

    return _has_header(wb, excel_path, select)
